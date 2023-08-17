

import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from openpyxl.workbook import Workbook

# Initialize Firebase
cred = credentials.Certificate("#######")
firebase_admin.initialize_app(cred, {
    'databaseURL': '##########'
})

# Specify the file path
file_path = 'FaceRecoAttendance.xlsx'

# Check if the file already exists
if os.path.exists(file_path):
    # Load the existing workbook
    workbook = load_workbook(file_path)
else:
    # Create a new workbook
    workbook = Workbook()

# Generate a sheet name based on the current date
current_date = datetime.now().strftime("%Y-%m-%d")
sub_name = input("Enter Subject Name :")
sheet_name = f"Attendance_{sub_name}"

# Check if the sheet already exists
if sheet_name in workbook.sheetnames:
    # If the sheet exists, delete it
    workbook.remove(workbook[sheet_name])

# Create a new sheet
worksheet = workbook.create_sheet(sheet_name)

# Retrieve data from Firebase
ref = db.reference('Students')  # Replace 'students' with the appropriate path in your database
data = ref.get()

# Write header row
header_row = ['Roll Number', 'Name', 'Date', '2 to 3', '3 to 4', '4 to 5', 'Total_Attendance']
worksheet.append(header_row)

# Set column widths and alignment
column_widths = [15, 20, 25, 15]
for i, column_width in enumerate(column_widths):
    column_letter = get_column_letter(i + 1)
    worksheet.column_dimensions[column_letter].width = column_width
    for cell in worksheet[column_letter]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Write data rows
for student_id, student_info in data.items():
    roll_no = student_info['roll_no']
    name = student_info['name']
    date_time = student_info['last_attendance_time']
    attendance1 = student_info['2 to 3']
    attendance2 = student_info['3 to 4']
    attendance3 = student_info['4 to 5']
    total_attendance = student_info['total_attendance']
    worksheet.append([roll_no, name, date_time, attendance1,attendance2, attendance3, total_attendance])

# Save the workbook
workbook.save(file_path)

